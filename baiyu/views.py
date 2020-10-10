from django.db import transaction
from django.shortcuts import render,HttpResponse,redirect,HttpResponseRedirect
from django.contrib import messages
from django.http import JsonResponse
from baiyu.db import *
from baiyu.function.introduced_total import *
from baiyu.function.zudai_to_fumudai import *
from baiyu.function.shangpindai import *
from baiyu.function.download import *
from baiyu.calc import *
import xlwt
from openpyxl import Workbook
import datetime
from openpyxl.styles import  Alignment,Font,colors
from io import BytesIO
import time
import xlrd
import logging
from login.views import check_login


# Create your views here.

num_progress = 0
logger = logging.getLogger('django')

@check_login
def index(request):
    return render(request, 'index.html')


'''
#渲染接口，展示祖代引种入舍总量
#by sujie 2019/05/27
'''
@check_login
def show_progenitor_introduced(request):
    content = {
        'introduced_info':get_introduced_detail_info(1,1),
        'count':get_count_ProgenitorIntroduced(1,1)
    }
    return render(request,'progenitor_introduced.html',content)


def zudai_statistics(request):
    bird_type = 1
    nGen = 1
    content  ={
        "zudai_statistic":get_zudai_statistic(bird_type,nGen),
        "statistic_count":get_daici_statistic_count(bird_type,nGen)
    }
    return render(request,'zudai_statistics.html',content)


def parent_statistics(request):
    bird_type = 1
    nGen = 2
    content = {
        "fumudai_statistic": get_zudai_statistic(bird_type, nGen),
        "statistic_count": get_daici_statistic_count(bird_type, nGen)
    }
    return render(request,'parent_statistics.html',content)

def shangpin_statistics(request):
    content = {
        'shangpin_statistics':get_sWeekly_statistic()
    }
    return render(request,'shangpin_statistics.html',content)

###################################################################
## written by sujie@boyar.cn   2019-05-27
##  Below is the standard paramter of calculate
###################################################################
'''
#渲染接口，展示祖代引种周度标准参数
#by sujie 2019/05/27
'''
def show_zWeekly_standard(request):
    bird_type = 1
    nGen = 1
    content = {
        'zhudai_weekly_standard':get_weekly_standard(bird_type,nGen)
    }
    return render(request,'zWeeklyStandard.html',content)


'''
#渲染接口，展示父母代周度标准参数
#by sujie 2019/05/27
'''
def show_fWeekly_standard(request):
    bird_type = 1
    nGen = 2
    content = {
        'fumudai_weekly_standard':get_weekly_standard(bird_type,nGen)
    }
    return render(request,'fWeeklyStandard.html',content)


'''
#渲染接口，展示商品代日度标准参数
#by sujie 2019/05/27
'''
def show_sDaily_standard(request):
    bird_type = 1
    content = {
        'shangpindai_daily_standard':get_shangpindai_daily_standard(bird_type)
    }
    return render(request,'sDailyStandard.html',content)

'''
#渲染接口，展示商品代年度参数
#by sujie 2019/05/27
'''
def show_sYearly_param(request):
    content = {
        'shangpindai_yearly_param':get_shangpindai_yearly_param()
    }
    return render(request,'sYearlyParam.html',content)


'''
#渲染接口，展示祖代总参数
#by sujie 2019/05/27
'''
def show_zudai_allParam(request):
    bird_type = 1
    nGen = 1
    content = {
        'zudai_whole_param':get_all_param(bird_type,nGen) #1为祖代，2为父母代
    }
    return render(request,'show_zd_allParam.html',content)

'''
#渲染接口，展示父母代总参数
#by sujie 2019/05/30
'''
def show_fumudai_allParam(request):
    bird_type = 1
    nGen = 2
    content = {
        'fmdai_whole_param':get_all_param(bird_type,nGen) #1为祖代，2为父母代
    }
    return render(request,'show_fmdai_allParam.html',content)


'''
#渲染接口，展示祖代年淘汰鸡肉参数
#by sujie 2019/05/27
'''
def zdYearly_TtJirou_Param(request):
    bird_type = 1
    nGen = 1
    content = {
        'zudai_tcjirou_param':get_tcjirou_param(bird_type,nGen) #1为祖代，2为父母代
    }
    return render(request,'zdYearly_TtJirouParam.html',content)

'''
#渲染接口，展示父母代年淘汰鸡肉参数
#by sujie 2019/05/27
'''
def fmdYearly_TtJirou_Param(request):
    bird_type = 1
    nGen = 2
    content = {
        'fmdai_tcjirou_param':get_tcjirou_param(bird_type,nGen) #1为祖代，2为父母代
    }
    return render(request,'fmdYearly_TtJirouParam.html',content)


def add_whole_param(request):
    return render(request,'add_whole_param.html')


'''
#渲染接口，展示周度标准日期
'''
def show_weeklydate_standard(request):
    content = {
        'weekly_standard_date':get_weeklydate_standard()
    }
    return render(request,'weekly_standard_date.html',content)


def show_company_info(request):
    content = {
        'company_info':get_company_info()
    }

    return render(request,'company_info.html',content)


def show_feedway_info(request):
    content = {
        'feedway_info':get_feedway_info()
    }

    return render(request,'feedway.html',content)

def show_species_info(request):
    content = {
        'species_info':get_species_info()
    }
    return render(request,'species_info.html',content)

def add_company(request):
    error_msg = ""
    if request.method == "POST":
        companyName = request.POST.get("companyName", None)  # print(new_name)
        Remark = request.POST.get('remark',None)
        CompanyInfo.objects.create(companyName=companyName,Remark=Remark)  # 数据库中新创建一条数据行
        return redirect("/baiyu/show_company_info/")  # redirect返回方法 HttpResponse返回字符串
    else:
        error_msg = "公司名称不能为空"
    return render(request,'company_add.html',{"error": error_msg})



def edit_company(request,company_id):
    if request.method == "GET":
        print('%'*20)
        company_info = CompanyInfo.objects.filter(id=company_id).first()
        return render(request, 'company_edit.html', {'company_info': company_info})
    elif request.method == "POST":
        company_name = request.POST.get("companyName")
        remark = request.POST.get("remark")
        CompanyInfo.objects.filter(id = company_id).update(companyName=company_name,Remark=remark)
        return redirect("/baiyu/show_company_info/")

def del_company(request,company_id):
    if request.method == "GET":
        try:
            CompanyInfo.objects.filter(id = company_id).delete()
        except Exception as e:
            print('e')
    return redirect("/baiyu/show_company_info/")


def add_qzhy(request,input_id):
    error_msg = ""
    if request.method == "GET":
        print('%'*20)
        qzhy_info = IntroducedInfoDetail.objects.filter(id=input_id).first()
        return render(request, 'qzhy_add.html', {'qzhy_info': qzhy_info})
    elif request.method == "POST":
        qzhyFlag_switch = request.POST.get("qzhyflag", None)  # print(new_name)
        qzhyRate = request.POST.get("qzhyRate", None)  # print(new_name)
        qzhyStartWeek = request.POST.get("qzhyStartWeek", None)  # print(new_name)
        HuanyuInterval = request.POST.get("HuanyuInterval",None)
        qzhyTime = request.POST.get("qzhyTime", None)  # print(new_name)
        Remark = request.POST.get('remark',None)
        if qzhyFlag_switch == 'on':
            qzhyFlag = 1
        else:
            qzhyFlag = 0
        print(qzhyFlag,qzhyRate,qzhyStartWeek,qzhyTime,Remark)
        IntroducedInfoDetail.objects.filter(id=input_id).update(
            qzhyFlag = qzhyFlag,
            qzhyStartWeek = qzhyStartWeek,
            huanyuRate = qzhyRate,
            HuanyuInterval = HuanyuInterval,
            qzhyPeriod = qzhyTime,
            Remark = Remark
        )
        return redirect("/baiyu/show_progenitor_introduced/")  # redirect返回方法 HttpResponse返回字符串
    else:
        error_msg = "强制换羽失败！"
    return render(request,'qzhy_add.html',{"error": error_msg})



def add_species(request):
    error_msg = ""
    if request.method == "POST":
        print(request.POST)
        SpeciesName = request.POST.get('SpeciesName',None)
        BirdType = request.POST.get('BirdType',None)
        Generation = request.POST.get('Generation',None)
        Remark = request.POST.get('Remark', None)
        print(SpeciesName,BirdType,Generation)
        # companyName = request.POST.get("companyName", None)  # print(new_name)
        # Remark = request.POST.get('remark',None)
        SpeciesInfo.objects.create(SpeciesName=SpeciesName,BirdType=BirdType,Generation=Generation,Remark=Remark)  # 数据库中新创建一条数据行
        return redirect("/baiyu/show_species_info/")  # redirect返回方法 HttpResponse返回字符串
    else:
        error_msg = "公司名称不能为空"
    return render(request,'species_add.html',{"error": error_msg})


def edit_species(request,species_id):
    if request.method == "GET":
        print('%'*20)
        species_info = SpeciesInfo.objects.filter(id=species_id).first()
        return render(request, 'species_edit.html', {'species_info': species_info})
    elif request.method == "POST":
        species_name = request.POST.get("SpeciesName")
        bird_type = request.POST.get("BirdType")
        generation = request.POST.get("Generation")
        remark = request.POST.get("Remark")
        SpeciesInfo.objects.filter(id = species_id).update(
            SpeciesName=species_name,
            BirdType = bird_type,
            Generation = generation,
            Remark=remark
        )
        return redirect("/baiyu/show_species_info/")


def del_species(request,species_id):
    if request.method == "GET":
        try:
            SpeciesInfo.objects.filter(id = species_id).delete()
        except Exception as e:
            print('e')
    return redirect("/baiyu/show_species_info/")


def add_feedway(request):
    error_msg = ""
    if request.method == "POST":
        feedWayName = request.POST.get("feedWayName", None)  # print(new_name)
        description = request.POST.get("description",None)
        remark = request.POST.get('remark',None)
        FeedWay.objects.create(feedWayName=feedWayName,description = description,remark=remark)  # 数据库中新创建一条数据行
        return redirect("/baiyu/show_feedway_info/")  # redirect返回方法 HttpResponse返回字符串
    else:
        error_msg = "公司名称不能为空"
    return render(request,'feedway_add.html',{"error": error_msg})


def edit_feedway(request,feedway_id):
    if request.method == "GET":
        print('%'*20)
        feedway_info = FeedWay.objects.filter(id=feedway_id).first()
        return render(request, 'feedway_edit.html', {'feedway_info': feedway_info})
    elif request.method == "POST":
        feedWayName = request.POST.get("feedWayName")
        description = request.POST.get("description")
        remark = request.POST.get("remark")
        FeedWay.objects.filter(id = feedway_id).update(
            feedWayName=feedWayName,
            description = description,
            remark=remark)
        return redirect("/baiyu/show_feedway_info/")


def del_feedway(request,feedway_id):
    if request.method == "GET":
        try:
            FeedWay.objects.filter(id = feedway_id).delete()
        except Exception as e:
            print('e')
    return redirect("/baiyu/show_feedway_info/")


'''
计算强制换羽
'''
def calc_qzhy(request):
    ### step1： 获取强制换羽的年份和周度


    ### step2： 删除原始时间是step1中年份的所有中间数据


    ### step3： 计算强制换羽的数据，正常数据和换羽数据
    return HttpResponse("Yes")


def introduced_total_info(request):
    flag = gen_total_introdeced_data()
    content = {
        'total_introduced_info': get_total_introduced_info()
    }
    return render(request,'total_introduced.html',content)

def show_total_introduced_info(request):
    bird_type = 1
    nGen = 1
    save_type = 1
    content  = {
        'total_introduced_info':get_total_introduced_info(bird_type,nGen,save_type)
    }
    return render(request, 'baiyu_total_introduced.html',content)


def start_calc_generation(request,bird_type,nGen,save_type):
    # print('HHHH'*20)
    # save_type = 1
    calc_generation_data(bird_type,nGen,save_type)
    messages.info(request, '计算完成！')
    return redirect('/baiyu/zudai_statistics/')


'''
展示界面 UI
'''
def show_progress1(request):
    # return JsonResponse(num_progress, safe=False)
    return render(request, 'calc_index.html')



'''
后台实际处理程序
'''
def process_data(request):
    # ...
    bird_type = 1
    nGen = 1
    global num_progress

    for i in range(123456):
        # ... 数据处理业务
        # calc_generation_data(bird_type, nGen)
        num_progress = i * 100 / 123456; # 更新后台进度值，因为想返回百分数所以乘100
        # print 'num_progress=' + str(num_progress)
        # time.sleep(1)
        res = int(num_progress)
        # print 'i='+str(i)
        print('res----=',str(res))
    return JsonResponse(res, safe=False)

'''
前端JS需要访问此程序来更新数据
'''
def show_progress(request):
    print('show_progress----------',str(num_progress))
    return JsonResponse(int(num_progress), safe=False)


def show_calc_index(request):
    return render(request,'calc_index.html')


def calc_shangpindai_statistics(request,bird_type,nGen,save_type):
    calc_shangpin_data(bird_type,nGen,save_type)
    return redirect('/baiyu/shangpin_statistics/')


# def export_excel_zudai_data(request):
#     res = export_excel()
#     return redirect('/baiyu/show_species_info')



# # 导出excel数据
# def export_excel(request):
#     # 设置HTTPResponse的类型
#     response = HttpResponse(content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = 'attachment;filename=order.xls'
#
#
#
#     # # 1. 创建一个Excel文件
#     # workbook = xlsxwriter.Workbook('demo1.xlsx')
#     #
#     # # 2. 创建一个工作表sheet对象
#     # worksheet = workbook.add_worksheet()
#
#     # 创建一个文件对象
#     wb = xlwt.Workbook(encoding='utf8')
#     # 创建一个sheet对象
#     sheet = wb.add_sheet('order-sheet')
#
#     # 设置文件头的样式,这个不是必须的可以根据自己的需求进行更改
#     style_heading = xlwt.easyxf("""
#             font:
#                 name Arial,
#                 colour_index white,
#                 bold on,
#                 height 0xA0;
#             align:
#                 wrap off,
#                 vert center,
#                 horiz center;
#             pattern:
#                 pattern solid,
#                 fore-colour 0x19;
#             borders:
#                 left THIN,
#                 right THIN,
#                 top THIN,
#                 bottom THIN;
#             """)
#
#     # 写入文件标题
#     sheet.write(0,0,'申请编号',style_heading)
#     sheet.write(0,1,'客户名称',style_heading)
#     sheet.write(0,2,'联系方式',style_heading)
#     sheet.write(0,3,'身份证号码',style_heading)
#     sheet.write(0,4,'办理日期',style_heading)
#     sheet.write(0,5,'处理人',style_heading)
#     sheet.write(0,6,'处理状态',style_heading)
#     sheet.write(0,7,'处理时间',style_heading)
#
#     # 写入数据
#     data_row = 1
#     # UserTable.objects.all()这个是查询条件,可以根据自己的实际需求做调整.
#     res_list = WeeklyCoreTable.objects.all().values().filter(nBirdsType=1,nGeneration=1)
#     print('%'*20,res_list)
#     for i in res_list:
#         print('####',i)
#         # 格式化datetime
#         # pri_time = i.pri_date.strftime('%Y-%m-%d')
#         # oper_time = i.operating_time.strftime('%Y-%m-%d')
#         sheet.write(data_row,0,i['Year'])
#         sheet.write(data_row,1,i['WeekNum'])
#         sheet.write(data_row,2,i['startDate'])
#         sheet.write(data_row,3,i['endDate'])
#         sheet.write(data_row,4,i['TotalYuChengCunLan'])
#         sheet.write(data_row,5,i['TotalChanDanCunLan'])
#         sheet.write(data_row,6,i['TotalDan'])
#         sheet.write(data_row,7,i['TotalChuJi'])
#         data_row = data_row + 1
#
#     # 写出到IO
#     output = BytesIO()
#     wb.save(output)
#     # 重新定位到开始
#     output.seek(0)
#     response.write(output.getvalue())
#     return response
#



def export_excel_zudai(request):
    wb = Workbook()  # optimized_write=True
    # 创建一个sheet
    sheet = wb.create_sheet('祖代_白羽_周度数据', 0)
    filename = 'zudai_baiyu_weekly_'+datetime.datetime.now().strftime('%Y-%m-%d_%H:%M:%S')
    table = wb.active
    #     # 写入文件标题
    table.cell(1,1,'年份')
    table.cell(1,2,'周度')
    table.cell(1,3,'开始时间')
    table.cell(1,4,'结束时间')
    table.cell(1,5,'育成期存栏')
    table.cell(1,6,'产蛋期存栏')
    table.cell(1,7,'产蛋数')
    table.cell(1,8,'一日龄雏鸡总数')
    table.cell(1, 9, '实际销售雏鸡总数')
    table.cell(1, 10, '淘汰鸡数目')
    table.cell(1, 11, '淘汰鸡肉量')
    table.cell(1, 12, '备注')
    # 写入数据
    data_row = 2
    res_list = WeeklyCoreTable.objects.all().values().filter(nBirdsType=1, nGeneration=1)
    print('%' * 20, res_list)
    for i in res_list:
        print('####', i)
        # 格式化datetime
        # pri_time = i.pri_date.strftime('%Y-%m-%d')
        # oper_time = i.operating_time.strftime('%Y-%m-%d')
        table.cell(data_row, 1, i['Year'])
        table.cell(data_row, 2, i['WeekNum'])
        table.cell(data_row, 3, i['startDate'])
        table.cell(data_row, 4, i['endDate'])
        table.cell(data_row, 5, i['TotalYuChengCunLan'])
        table.cell(data_row, 6, i['TotalChanDanCunLan'])
        table.cell(data_row, 7, i['TotalDan'])
        table.cell(data_row, 8, i['TotalChuJi'])
        table.cell(data_row, 9, i['TotalFactSaleChuJi'])
        table.cell(data_row, 10, i['TaoTaiJiNum'])
        table.cell(data_row, 11, i['dTaoTaiJiRou'])
        table.cell(data_row, 12, i['Remark'])
        data_row = data_row + 1

    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = 'attachment;filename=%s.xlsx' % filename
    wb.save(response)
    return response


def export_excel_fumudai(request):
    wb = Workbook()  # optimized_write=True
    # 创建一个sheet
    sheet = wb.create_sheet('父母代_白羽肉鸡_周度数据汇总', 0)
    filename = 'fumudai_baiyu_weekly_'+datetime.datetime.now().strftime('%Y-%m-%d_%H:%M:%S')
    # mm =
    # filename = 'sujie'
    table = wb.active
    #     # 写入文件标题
    table.cell(1,1,'年份')
    table.cell(1,2,'周度')
    table.cell(1,3,'开始时间')
    table.cell(1,4,'结束时间')
    table.cell(1,5,'育成期存栏')
    table.cell(1,6,'产蛋期存栏')
    table.cell(1,7,'产蛋数')
    table.cell(1,8,'一日龄雏鸡总数')
    table.cell(1, 9, '实际销售雏鸡总数')
    table.cell(1, 10, '淘汰鸡数目')
    table.cell(1, 11, '淘汰鸡肉量')
    table.cell(1, 12, '备注')
    # 写入数据
    data_row = 2
    res_list = WeeklyCoreTable.objects.all().values().filter(nBirdsType=1, nGeneration=2)
    print('%' * 20, res_list)
    for i in res_list:
        print('####', i)
        # 格式化datetime

        # oper_time = i.operating_time.strftime('%Y-%m-%d')
        table.cell(data_row, 1, i['Year'])
        table.cell(data_row, 2, i['WeekNum'])
        table.cell(data_row, 3, i['startDate'])
        table.cell(data_row, 4, i['endDate'])
        table.cell(data_row, 5, i['TotalYuChengCunLan'])
        table.cell(data_row, 6, i['TotalChanDanCunLan'])
        table.cell(data_row, 7, i['TotalDan'])
        table.cell(data_row, 8, i['TotalChuJi'])
        table.cell(data_row, 9, i['TotalFactSaleChuJi'])
        table.cell(data_row, 10, i['TaoTaiJiNum'])
        table.cell(data_row, 11, i['dTaoTaiJiRou'])
        table.cell(data_row, 12, i['Remark'])
        data_row = data_row + 1

    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = 'attachment;filename=%s.xlsx' % filename
    wb.save(response)
    return response



def export_excel_shangpindai(request):
    wb = Workbook()  # optimized_write=True
    # 创建一个sheet
    sheet = wb.create_sheet('商品代_白羽肉鸡_周度数据汇总', 0)
    filename = 'shangpindai_baiyu_weekly_'+datetime.datetime.now().strftime('%Y-%m-%d_%H:%M:%S')
    # mm =
    # filename = 'sujie'
    table = wb.active
    #     # 写入文件标题
    table.cell(1,1,'年份')
    table.cell(1,2,'周度')
    table.cell(1,3,'开始时间')
    table.cell(1,4,'结束时间')
    table.cell(1,5,'35日龄存栏量')
    table.cell(1,6,'42日龄存栏量')
    table.cell(1,7,'49日龄存栏量')
    table.cell(1,8,'56日龄存栏量')
    table.cell(1, 9, '35日龄出栏肉鸡数量')
    table.cell(1, 10, '42日龄出栏肉鸡数量')
    table.cell(1, 11, '49日龄出栏肉鸡数量')
    table.cell(1, 12, '56日龄出栏肉鸡数量')
    table.cell(1, 13, '出栏肉鸡总量')
    table.cell(1, 14, '35日龄出栏肉鸡活重')
    table.cell(1, 15, '42日龄出栏肉鸡活重')
    table.cell(1, 16, '49日龄出栏肉鸡活重')
    table.cell(1, 17, '56日龄出栏肉鸡活重')
    table.cell(1, 18, '出栏肉鸡活重总量')
    table.cell(1, 19, '35日龄出栏鸡肉重量')
    table.cell(1, 20, '42日龄出栏鸡肉重量')
    table.cell(1, 21, '49日龄出栏鸡肉重量')
    table.cell(1, 22, '56日龄出栏鸡肉重量')
    table.cell(1, 23, '出栏鸡肉总重量')
    table.cell(1, 24, '35日龄出栏鸡胸肉重量')
    table.cell(1, 25, '42日龄出栏鸡胸肉重量')
    table.cell(1, 26, '49日龄出栏鸡胸肉重量')
    table.cell(1, 27, '56日龄出栏鸡胸肉重量')
    table.cell(1, 28, '出栏鸡胸肉总重量')
    table.cell(1, 29, '35日龄出栏鸡翅重量')
    table.cell(1, 30, '42日龄出栏鸡翅重量')
    table.cell(1, 31, '49日龄出栏鸡翅重量')
    table.cell(1, 32, '56日龄出栏鸡翅重量')
    table.cell(1, 33, '出栏鸡翅总重量')
    table.cell(1, 34, '35日龄出栏鸡腿重量')
    table.cell(1, 35, '42日龄出栏鸡腿重量')
    table.cell(1, 36, '49日龄出栏鸡腿重量')
    table.cell(1, 37, '56日龄出栏鸡腿重量')
    table.cell(1, 38, '出栏鸡腿总重量')
    table.cell(1, 39, '35日龄出栏鸡骨架重量')
    table.cell(1, 40, '42日龄出栏鸡骨架重量')
    table.cell(1, 41, '49日龄出栏鸡骨架重量')
    table.cell(1, 42, '56日龄出栏鸡骨架重量')
    table.cell(1, 43, '出栏鸡骨架总重量')
    table.cell(1, 44, '35日龄出栏鸡内脏重量')
    table.cell(1, 45, '42日龄出栏鸡内脏重量')
    table.cell(1, 46, '49日龄出栏鸡内脏重量')
    table.cell(1, 47, '56日龄出栏鸡内脏重量')
    table.cell(1, 48, '出栏鸡内脏总重量')
    table.cell(1, 49, '备注')

    # 写入数据
    data_row = 2
    res_list = WeeklyStatisticTable.objects.all().values().filter(nBirdsType=1)
    print('%' * 20, res_list)
    for i in res_list:
        print('####', i)
        # 格式化datetime

        # oper_time = i.operating_time.strftime('%Y-%m-%d')
        table.cell(data_row, 1, i['Year'])
        table.cell(data_row, 2, i['WeekNum'])
        table.cell(data_row, 3, i['startDate'])
        table.cell(data_row, 4, i['endDate'])
        table.cell(data_row, 5, i['CunlLanNum35'])
        table.cell(data_row, 6, i['CunlLanNum42'])
        table.cell(data_row, 7, i['CunlLanNum49'])
        table.cell(data_row, 8, i['CunlLanNum56'])
        table.cell(data_row, 9, i['ChuLanRouJiNum35'])
        table.cell(data_row, 10, i['ChuLanRouJiNum42'])
        table.cell(data_row, 11, i['ChuLanRouJiNum49'])
        table.cell(data_row, 12, i['ChuLanRouJiNum56'])
        table.cell(data_row, 13, i['TotalChuLanRouJiNum'])
        table.cell(data_row, 14, i['HuoZhong35'])
        table.cell(data_row, 15, i['HuoZhong42'])
        table.cell(data_row, 16, i['HuoZhong49'])
        table.cell(data_row, 17, i['HuoZhong56'])
        table.cell(data_row, 18, i['TotalHuoZhong'])
        table.cell(data_row, 19, i['JiRou35'])
        table.cell(data_row, 20, i['JiRou42'])
        table.cell(data_row, 21, i['JiRou49'])
        table.cell(data_row, 22, i['JiRou56'])
        table.cell(data_row, 23, i['TotalJiRou'])
        table.cell(data_row, 24, i['JiXiong35'])
        table.cell(data_row, 25, i['JiXiong42'])
        table.cell(data_row, 26, i['JiXiong49'])
        table.cell(data_row, 27, i['JiXiong56'])
        table.cell(data_row, 28, i['TotalJiXiong'])
        table.cell(data_row, 29, i['JiChi35'])
        table.cell(data_row, 30, i['JiChi42'])
        table.cell(data_row, 31, i['JiChi49'])
        table.cell(data_row, 32, i['JiChi56'])
        table.cell(data_row, 33, i['TotalJiChi'])
        table.cell(data_row, 34, i['JiTui35'])
        table.cell(data_row, 35, i['JiTui42'])
        table.cell(data_row, 36, i['JiTui49'])
        table.cell(data_row, 37, i['JiTui56'])
        table.cell(data_row, 38, i['TotalJiTui'])
        table.cell(data_row, 39, i['JiGuJia35'])
        table.cell(data_row, 40, i['JiGuJia42'])
        table.cell(data_row, 41, i['JiGuJia49'])
        table.cell(data_row, 42, i['JiGuJia56'])
        table.cell(data_row, 43, i['TotalJiGuJia'])
        table.cell(data_row, 44, i['JiNeiZang35'])
        table.cell(data_row, 45, i['JiNeiZang42'])
        table.cell(data_row, 46, i['JiNeiZang49'])
        table.cell(data_row, 47, i['JiNeiZang56'])
        table.cell(data_row, 48, i['TotalJiNeiZang'])
        table.cell(data_row, 49, i['Remark'])
        data_row = data_row + 1

    response = HttpResponse(content_type='application/msexcel')
    response['Content-Disposition'] = 'attachment;filename=%s.xlsx' % filename
    wb.save(response)
    return response



def uploadZudaiIntroduced(request):
    '''
    祖代引种信息导入
    :param request:
    :return:
    '''
    if request.method == 'POST':
        f = request.FILES.get('zd_introduced')
        print(request)
        print(request.POST)
        print(request.FILES)
        excel_type = f.name.split('.')[1]
        if excel_type in ['xlsx','xls']:
            # 开始解析上传的excel表格
            wb = xlrd.open_workbook(filename=None,file_contents=f.read())
            table = wb.sheets()[0]
            rows = table.nrows  # 总行数
            try:
                with transaction.atomic():  # 控制数据库事务交易
                    for i in range(1,rows):
                        print('#'*20)
                        rowVlaues = table.row_values(i)
                        year = int(rowVlaues[0])
                        week = int(rowVlaues[1])
                        startDate,endDate = get_start_end_date_by_week(year,week)
                        companyName = rowVlaues[2]
                        speciesName = rowVlaues[3]
                        feedwayName = rowVlaues[4]
                        print(rowVlaues)
                        company_id = CompanyInfo.objects.values('id').filter(companyName=companyName).first()['id']
                        species_id = SpeciesInfo.objects.values('id').filter(SpeciesName=speciesName).first()['id']
                        feedway_id = FeedWay.objects.values('id').filter(feedWayName=feedwayName).first()['id']
                        print(
                            year,
                            week,
                            startDate,
                            endDate,
                            company_id,
                            species_id,
                            feedway_id,
                            int(rowVlaues[5]),
                            int(rowVlaues[6])
                        )
                        IntroducedInfoDetail.objects.create(
                            Year = year,
                            WeekNum = week,
                            startDate = startDate,
                            endDate = endDate,
                            CompanyId = company_id,
                            SpeciesId = species_id,
                            feedWayId = feedway_id,
                            RuSheNum = int(rowVlaues[5]),
                            LivePeriod = int(rowVlaues[6]),
                            nGeneration = 1,
                            nDraftOrOriginal=1,
                            nBirdsType=1
                        )
            except:
                logger.error('解析excel文件或者数据插入错误')
            return render(request,'date_error.html',{'message':'导入成功'})
        else:
            logger.error('上传文件类型错误！')
            return render(request,'date_error.html',{'message':'导入失败'})


'''
接受前台传来的excel
'''
def import_origin_introduced(request):
    '''
    祖代引种信息导入
    :param request:
    :return:
    '''
    if request.method == 'POST':
        f = request.FILES.get('zd_introduced_total')
        print(request)
        print(request.POST)
        print(request.FILES)
        excel_type = f.name.split('.')[1]
        if excel_type in ['xlsx','xls']:
            # 开始解析上传的excel表格
            wb = xlrd.open_workbook(filename=None,file_contents=f.read())
            table = wb.sheets()[0]
            rows = table.nrows  # 总行数
            try:
                with transaction.atomic():  # 控制数据库事务交易
                    for i in range(1,rows):
                        print('#'*20)
                        rowVlaues = table.row_values(i)
                        year = int(rowVlaues[0])
                        week = int(rowVlaues[1])
                        startDate,endDate = get_start_end_date_by_week(year,week)
                        print(rowVlaues)
                        print(
                            year,
                            week,
                            startDate,
                            endDate,
                            int(rowVlaues[2]),
                            int(rowVlaues[3])
                        )
                        IntroducedInfo.objects.create(
                            Year = year,
                            WeekNum = week,
                            startDate = startDate,
                            endDate = endDate,
                            RuSheNum = int(rowVlaues[2]),
                            LivePeriod = int(rowVlaues[3]),
                            nGeneration = 1,
                            nDraftOrOriginal = 1,
                            nBirdsType = 1,
                            flag=0
                        )
                        IntroducedInfo.objects.create(
                            Year=year,
                            WeekNum=week,
                            startDate=startDate,
                            endDate=endDate,
                            RuSheNum=int(rowVlaues[2]),
                            LivePeriod=int(rowVlaues[3]),
                            nGeneration=1,
                            nDraftOrOriginal=2,
                            nBirdsType=1,
                            flag=0
                        )
            except:
                logger.error('解析excel文件或者数据插入错误')
            messages.success(request,'导入成功')
            return HttpResponseRedirect('/baiyu/show_total_introduced_info/')
        else:
            messages.error(request,'上传文件类型错误！')
            return HttpResponseRedirect('/baiyu/show_total_introduced_info/')


def zdjiaozheng(request):
    # year_list = [x for x in range(2000,2050)]
    content = {
        "year_list":[x for x in range(2000,2051)],
        "week_list":[x for x in range(1,54)],
        "sitaorate_info":get_sitaorate_correct_info(1),
        "chandanrate_info":get_chandanrate_correct_info(1),
        "rufurate_info":get_rufurate_correct_info(1),
        "shoujingrate_info":get_shoujingrate_correct_info(1),
        "fuhuarate_info": get_fuhuarate_correct_info(1),
        "jianchurate_info": get_jianchurate_correct_info(1),
        "salerate_info": get_salerate_correct_info(1),
    }
    return render(request, 'baiyu/zudai_jiaozheng.html',content)

def zudai_correct_sitaorate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_1')
        start_week = request.POST.get('start_week_1')
        end_year = request.POST.get('end_year_1')
        end_week = request.POST.get('end_week_1')
        correct_sitaorate = request.POST.get('correct_sitaorate')
        sitaorate_notes = request.POST.get('sitaorate_notes')
        print(start_year,start_week,end_year,end_week,correct_sitaorate,sitaorate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                SiTaoC = correct_sitaorate,
                SiTaoCNote = sitaorate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")



def zudai_correct_chandanrate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_2')
        start_week = request.POST.get('start_week_2')
        end_year = request.POST.get('end_year_2')
        end_week = request.POST.get('end_week_2')
        correct_chandanrate = request.POST.get('correct_chandanrate')
        chandanrate_notes = request.POST.get('chandanrate_notes')
        print(start_year,start_week,end_year,end_week,correct_chandanrate,chandanrate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                ChanDanC = correct_chandanrate,
                ChanDanCNote = chandanrate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")


def zudai_correct_rufurate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_3')
        start_week = request.POST.get('start_week_3')
        end_year = request.POST.get('end_year_3')
        end_week = request.POST.get('end_week_3')
        correct_rufurate = request.POST.get('correct_rufurate')
        rufurate_notes = request.POST.get('rufurate_notes')
        print(start_year,start_week,end_year,end_week,correct_rufurate,rufurate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                RuFuZhongDanC = correct_rufurate,
                RuFuZhongDanCNote = rufurate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")

def zudai_correct_shoujingrate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_4')
        start_week = request.POST.get('start_week_4')
        end_year = request.POST.get('end_year_4')
        end_week = request.POST.get('end_week_4')
        correct_shoujingrate = request.POST.get('correct_shoujingrate')
        shoujingrate_notes = request.POST.get('shoujingrate_notes')
        print(start_year,start_week,end_year,end_week,correct_shoujingrate,shoujingrate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                ShouJingC = correct_shoujingrate,
                ShouJingCNote = shoujingrate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")

def zudai_correct_fuhuarate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_5')
        start_week = request.POST.get('start_week_5')
        end_year = request.POST.get('end_year_5')
        end_week = request.POST.get('end_week_5')
        correct_fuhuarate = request.POST.get('correct_fuhuarate')
        fuhuarate_notes = request.POST.get('fuhuarate_notes')
        print(start_year,start_week,end_year,end_week,correct_fuhuarate,fuhuarate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                FuHuaC = correct_fuhuarate,
                FuHuaCNote = fuhuarate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")

def zudai_correct_jianchurate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_6')
        start_week = request.POST.get('start_week_6')
        end_year = request.POST.get('end_year_6')
        end_week = request.POST.get('end_week_6')
        correct_jianchurate = request.POST.get('correct_jianchurate')
        jianchurate_notes = request.POST.get('jianchurate_notes')
        print(start_year,start_week,end_year,end_week,correct_jianchurate,jianchurate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                JianChuC = correct_jianchurate,
                JianChuCNote = jianchurate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")

def zudai_correct_salerate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_7')
        start_week = request.POST.get('start_week_7')
        end_year = request.POST.get('end_year_7')
        end_week = request.POST.get('end_week_7')
        correct_salerate = request.POST.get('correct_salerate')
        salerate_notes = request.POST.get('salerate_notes')
        print(start_year,start_week,end_year,end_week,correct_salerate,salerate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,1):
            WeekCorrectionFactor.objects.filter(nGeneration=1).filter(id = index).update(
                SaleRateC = correct_salerate,
                SaleRateCNote = salerate_notes
            )
        return redirect("/baiyu/zdjiaozheng/")


def fmdjiaozheng(request):
    # year_list = [x for x in range(2000,2050)]
    content = {
        "year_list":[x for x in range(2000,2051)],
        "week_list":[x for x in range(1,54)],
        "sitaorate_info":get_sitaorate_correct_info(2),
        "chandanrate_info":get_chandanrate_correct_info(2),
        "rufurate_info":get_rufurate_correct_info(2),
        "shoujingrate_info":get_shoujingrate_correct_info(2),
        "fuhuarate_info": get_fuhuarate_correct_info(2),
        "jianchurate_info": get_jianchurate_correct_info(2),
        "salerate_info": get_salerate_correct_info(2),
    }
    return render(request, 'baiyu/fmdai_jiaozheng.html',content)

def fmdai_correct_sitaorate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_1')
        start_week = request.POST.get('start_week_1')
        end_year = request.POST.get('end_year_1')
        end_week = request.POST.get('end_week_1')
        correct_sitaorate = request.POST.get('correct_sitaorate')
        sitaorate_notes = request.POST.get('sitaorate_notes')
        print(start_year,start_week,end_year,end_week,correct_sitaorate,sitaorate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                SiTaoC = correct_sitaorate,
                SiTaoCNote = sitaorate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")



def fmdai_correct_chandanrate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_2')
        start_week = request.POST.get('start_week_2')
        end_year = request.POST.get('end_year_2')
        end_week = request.POST.get('end_week_2')
        correct_chandanrate = request.POST.get('correct_chandanrate')
        chandanrate_notes = request.POST.get('chandanrate_notes')
        print(start_year,start_week,end_year,end_week,correct_chandanrate,chandanrate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                ChanDanC = correct_chandanrate,
                ChanDanCNote = chandanrate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")


def fmdai_correct_rufurate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_3')
        start_week = request.POST.get('start_week_3')
        end_year = request.POST.get('end_year_3')
        end_week = request.POST.get('end_week_3')
        correct_rufurate = request.POST.get('correct_rufurate')
        rufurate_notes = request.POST.get('rufurate_notes')
        print(start_year,start_week,end_year,end_week,correct_rufurate,rufurate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                RuFuZhongDanC = correct_rufurate,
                RuFuZhongDanCNote = rufurate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")

def fmdai_correct_shoujingrate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_4')
        start_week = request.POST.get('start_week_4')
        end_year = request.POST.get('end_year_4')
        end_week = request.POST.get('end_week_4')
        correct_shoujingrate = request.POST.get('correct_shoujingrate')
        shoujingrate_notes = request.POST.get('shoujingrate_notes')
        print(start_year,start_week,end_year,end_week,correct_shoujingrate,shoujingrate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                ShouJingC = correct_shoujingrate,
                ShouJingCNote = shoujingrate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")

def fmdai_correct_fuhuarate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_5')
        start_week = request.POST.get('start_week_5')
        end_year = request.POST.get('end_year_5')
        end_week = request.POST.get('end_week_5')
        correct_fuhuarate = request.POST.get('correct_fuhuarate')
        fuhuarate_notes = request.POST.get('fuhuarate_notes')
        print(start_year,start_week,end_year,end_week,correct_fuhuarate,fuhuarate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                FuHuaC = correct_fuhuarate,
                FuHuaCNote = fuhuarate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")

def fmdai_correct_jianchurate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_6')
        start_week = request.POST.get('start_week_6')
        end_year = request.POST.get('end_year_6')
        end_week = request.POST.get('end_week_6')
        correct_jianchurate = request.POST.get('correct_jianchurate')
        jianchurate_notes = request.POST.get('jianchurate_notes')
        print(start_year,start_week,end_year,end_week,correct_jianchurate,jianchurate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                JianChuC = correct_jianchurate,
                JianChuCNote = jianchurate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")

def fmdai_correct_salerate(request):
    if request.method == 'POST':
        print(request.POST)
        start_year = request.POST.get('start_year_7')
        start_week = request.POST.get('start_week_7')
        end_year = request.POST.get('end_year_7')
        end_week = request.POST.get('end_week_7')
        correct_salerate = request.POST.get('correct_salerate')
        salerate_notes = request.POST.get('salerate_notes')
        print(start_year,start_week,end_year,end_week,correct_salerate,salerate_notes)
        for index in get_date_id_list(start_year,start_week,end_year,end_week,2):
            WeekCorrectionFactor.objects.filter(nGeneration=2).filter(id = index).update(
                SaleRateC = correct_salerate,
                SaleRateCNote = salerate_notes
            )
        return redirect("/baiyu/fmdjiaozheng/")


def tongji_detail(request):
    '''
    统计固定年份下的某个公司饲养的某个品种的数量
    :param request:
    :return:
    '''
    print('*'*20)
    print(request.POST)
    if request.method == 'POST':
        print('^^^'*10)
        year = request.POST.get("year")
        week = request.POST.get("week")

    else:
        year = 2019
        week = 20
    nGen = 1
    res_list = []
    speciesID_list = [x['id'] for x in SpeciesInfo.objects.all().values('id').order_by('id').distinct()]
    cpmpanyID_list = [x['id'] for x in CompanyInfo.objects.all().values('id').order_by('id').distinct()]
    for company in cpmpanyID_list:
        tmp = []
        tmp.append(CompanyInfo.objects.values('companyName').filter(id=company).first()['companyName'])
        for species in speciesID_list:
            tmp_num = get_tongji_data_by_company_and_species(year, week, nGen, company, species)
            tmp.append(tmp_num)
        tmp.append(list_add(tmp[1:]))
        res_list.append(tmp)
        del tmp

    total_by_species = []
    total_by_species.append('总计')
    for index in speciesID_list:
        species_total_num = get_tongji_data_by_species(year, week, nGen, index)
        total_by_species.append(species_total_num)
    total_by_species.append(list_add(total_by_species[1:]))
    content = {
        "year_list": [x for x in range(2000, 2051)],
        "week_list": [x for x in range(1, 54)],
        'species_list':[x['SpeciesName'] for x in SpeciesInfo.objects.all().values('SpeciesName').order_by('id')],
        'res_list':res_list,
        'total_by_species':total_by_species,
        'show_year':year,
        'show_week':week
    }
    return render(request,'tongji_detail.html',content)


def yearly_tongji_detail(request):
    '''
    统计固定年份下的某个公司饲养的某个品种的数量
    :param request:
    :return:
    '''
    print('*'*20)
    print(request.POST)
    if request.method == 'POST':
        print('^^^'*10)
        year = request.POST.get("year")
        # week = request.POST.get("week")

    else:
        year = 2019
        # week = 20
    nGen = 1
    res_list = []
    speciesID_list = [x['id'] for x in SpeciesInfo.objects.all().values('id').order_by('id').distinct()]
    cpmpanyID_list = [x['id'] for x in CompanyInfo.objects.all().values('id').order_by('id').distinct()]
    for company in cpmpanyID_list:
        tmp = []
        tmp.append(CompanyInfo.objects.values('companyName').filter(id=company).first()['companyName'])
        for species in speciesID_list:
            tmp_num = get_yearly_tongji_data_by_company_and_species(year, nGen, company, species)
            tmp.append(tmp_num)
        tmp.append(list_add(tmp[1:]))
        res_list.append(tmp)
        del tmp

    total_by_species = []
    total_by_species.append('总计')
    for index in speciesID_list:
        species_total_num = get_yearly_tongji_data_by_species(year, nGen, index)
        total_by_species.append(species_total_num)
    total_by_species.append(list_add(total_by_species[1:]))
    content = {
        "year_list": [x for x in range(2000, 2051)],
        'species_list':[x['SpeciesName'] for x in SpeciesInfo.objects.all().values('SpeciesName').order_by('id')],
        'res_list':res_list,
        'total_by_species':total_by_species,
        'show_year':year
    }
    return render(request,'yearly_tongji_detail.html',content)